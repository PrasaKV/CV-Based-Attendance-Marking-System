import os
import re
import csv
import xml.etree.ElementTree as ET
import json
import io
import cv2
import numpy as np
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class CVEngine:
    """Enhanced Computer Vision Engine (v2.1) with Auto-Deskew, Morphological Table Line
    Suppression, Multi-format Roster Parsing (XML, JSON, CSV, XLSX), Face Detection, and
    Confidence Scoring.
    """

    def __init__(self, upload_folder, config=None):
        self.upload_folder = upload_folder
        self.config = config
        
        # Load Haar Cascade Face Classifier safely
        self.face_cascade = None
        try:
            if hasattr(cv2, 'CascadeClassifier') and hasattr(cv2, 'data'):
                cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
                if os.path.exists(cascade_path):
                    self.face_cascade = cv2.CascadeClassifier(cascade_path)
        except Exception as e:
            print(f"[CVEngine] Face cascade initialization warning: {e}")

    # =========================================================================
    # Multi-Format Roster Parsers (XML, JSON, CSV, XLSX)
    # =========================================================================

    def parse_student_file(self, file_path):
        """Parse student index and name from XML, JSON, CSV, or Excel files"""
        students = []
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".xml":
            students = self._parse_xml(file_path)
        elif ext == ".json":
            students = self._parse_json(file_path)
        elif ext == ".csv":
            students = self._parse_csv(file_path)
        elif ext in [".xlsx", ".xls"]:
            students = self._parse_excel(file_path)

        return students

    def _parse_xml(self, xml_file):
        """Parse student records from structured or raw XML roster"""
        students = []
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()

            for student_elem in root.iter("student"):
                idx_elem = student_elem.find("index")
                name_elem = student_elem.find("name")
                if idx_elem is not None and name_elem is not None:
                    students.append({
                        "index": idx_elem.text.strip() if idx_elem.text else "",
                        "name": name_elem.text.strip() if name_elem.text else ""
                    })

            if not students:
                with open(xml_file, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()

                indices = re.findall(r"<index[^>]*>([^<]+)</index>", content)
                names = re.findall(r"<name>([^<]+)</name>", content)

                for idx, name in zip(indices, names):
                    students.append({
                        "index": idx.strip(),
                        "name": name.strip()
                    })

        except Exception as e:
            print(f"[CVEngine] XML Parsing error: {e}")
            try:
                with open(xml_file, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
                indices = re.findall(r"<index[^>]*>([^<]+)</index>", content)
                names = re.findall(r"<name>([^<]+)</name>", content)
                for idx, name in zip(indices, names):
                    students.append({"index": idx.strip(), "name": name.strip()})
            except Exception as inner_e:
                print(f"[CVEngine] Fallback XML parse failed: {inner_e}")

        return students

    def _parse_json(self, json_file):
        """Parse student records from JSON roster file"""
        students = []
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            if isinstance(data, list):
                for item in data:
                    students.append({
                        "index": str(item.get("index", "")).strip(),
                        "name": str(item.get("name", "")).strip()
                    })
            elif isinstance(data, dict) and "students" in data:
                for item in data["students"]:
                    students.append({
                        "index": str(item.get("index", "")).strip(),
                        "name": str(item.get("name", "")).strip()
                    })
        except Exception as e:
            print(f"[CVEngine] JSON Parsing error: {e}")

        return students

    def _parse_csv(self, csv_file):
        """Parse student records from CSV roster with automatic header detection"""
        students = []
        try:
            with open(csv_file, "r", encoding="utf-8-sig", errors="ignore") as f:
                # Detect delimiter (comma, tab, semicolon)
                sample = f.read(2048)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                    delimiter = dialect.delimiter
                except Exception:
                    delimiter = ","

                reader = csv.reader(f, delimiter=delimiter)
                rows = [row for row in reader if any(cell.strip() for cell in row)]

            if not rows:
                return []

            # Check if first row contains headers
            header = [c.strip().lower() for c in rows[0]]
            idx_col, name_col = -1, -1

            index_aliases = {"index", "index_no", "index_number", "student_index", "id", "reg_no", "regno", "student_id"}
            name_aliases = {"name", "student_name", "full_name", "fullname", "student"}

            for col_idx, col_name in enumerate(header):
                cleaned = re.sub(r'[^a-z0-9_]', '', col_name)
                if idx_col == -1 and cleaned in index_aliases:
                    idx_col = col_idx
                elif name_col == -1 and cleaned in name_aliases:
                    name_col = col_idx

            # Default to column 0 (index) and column 1 (name) if header not recognized
            data_rows = rows[1:] if (idx_col != -1 or name_col != -1) else rows
            if idx_col == -1:
                idx_col = 0
            if name_col == -1:
                name_col = 1 if len(rows[0]) > 1 else 0

            for row in data_rows:
                if len(row) > idx_col:
                    s_idx = row[idx_col].strip()
                    s_name = row[name_col].strip() if len(row) > name_col else ""
                    if s_idx:
                        students.append({"index": s_idx, "name": s_name})

        except Exception as e:
            print(f"[CVEngine] CSV Parsing error: {e}")

        return students

    def _parse_excel(self, xlsx_file):
        """Parse student records from Excel (.xlsx/.xls) roster"""
        students = []
        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                return []

            # Filter non-empty rows
            non_empty_rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
            if not non_empty_rows:
                return []

            # Find header
            first_row = [str(c).strip().lower() if c is not None else "" for c in non_empty_rows[0]]
            idx_col, name_col = -1, -1

            index_aliases = {"index", "index_no", "index_number", "student_index", "id", "reg_no", "regno", "student_id"}
            name_aliases = {"name", "student_name", "full_name", "fullname", "student"}

            for col_idx, col_val in enumerate(first_row):
                cleaned = re.sub(r'[^a-z0-9_]', '', col_val)
                if idx_col == -1 and cleaned in index_aliases:
                    idx_col = col_idx
                elif name_col == -1 and cleaned in name_aliases:
                    name_col = col_idx

            data_rows = non_empty_rows[1:] if (idx_col != -1 or name_col != -1) else non_empty_rows
            if idx_col == -1:
                idx_col = 0
            if name_col == -1:
                name_col = 1 if len(non_empty_rows[0]) > 1 else 0

            for row in data_rows:
                if len(row) > idx_col and row[idx_col] is not None:
                    s_idx = str(row[idx_col]).strip()
                    s_name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] is not None else ""
                    if s_idx:
                        students.append({"index": s_idx, "name": s_name})

        except Exception as e:
            print(f"[CVEngine] Excel Parsing error: {e}")

        return students

    # =========================================================================
    # Computer Vision Enhancements (Deskew & Morphological Line Filtering)
    # =========================================================================

    def _deskew_image(self, image, gray_img):
        """Detects skew angle in scanned attendance sheets and rotates image to be upright"""
        try:
            # Otsu thresholding to get document contours
            _, thresh = cv2.threshold(gray_img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
            coords = np.column_stack(np.where(thresh > 0))
            if len(coords) < 100:
                return image, gray_img, 0.0

            angle = cv2.minAreaRect(coords)[-1]
            if angle < -45:
                angle = -(90 + angle)
            elif angle > 45:
                angle = 90 - angle
            else:
                angle = -angle

            # Apply rotation only if tilt is between 0.5 and 20 degrees
            if 0.5 <= abs(angle) <= 20.0:
                (h, w) = gray_img.shape[:2]
                center = (w // 2, h // 2)
                M = cv2.getRotationMatrix2D(center, angle, 1.0)
                deskewed_color = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
                deskewed_gray = cv2.warpAffine(gray_img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
                return deskewed_color, deskewed_gray, round(angle, 2)
        except Exception as e:
            print(f"[CVEngine] Deskew notice: {e}")

        return image, gray_img, 0.0

    def _clean_signature_roi(self, thresh_roi):
        """Morphologically detects and suppresses horizontal/vertical table grid lines
        and small scanner noise specks from the signature ROI.
        """
        roi_h, roi_w = thresh_roi.shape
        if roi_h < 4 or roi_w < 4:
            return thresh_roi, int(cv2.countNonZero(thresh_roi))

        # 1. Detect horizontal table border lines
        horiz_size = max(10, roi_w // 3)
        horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (horiz_size, 1))
        horiz_lines = cv2.morphologyEx(thresh_roi, cv2.MORPH_OPEN, horiz_kernel)

        # 2. Detect vertical table border lines
        vert_size = max(8, roi_h // 2)
        vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, vert_size))
        vert_lines = cv2.morphologyEx(thresh_roi, cv2.MORPH_OPEN, vert_kernel)

        # Combine detected grid lines
        table_grid = cv2.bitwise_or(horiz_lines, vert_lines)

        # Subtract grid lines from signature ROI
        cleaned_roi = cv2.subtract(thresh_roi, table_grid)

        # 3. Filter out isolated noise specks (< 2x2)
        noise_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        cleaned_roi = cv2.morphologyEx(cleaned_roi, cv2.MORPH_OPEN, noise_kernel)

        stroke_pixels = int(cv2.countNonZero(cleaned_roi))
        return cleaned_roi, stroke_pixels

    def _calculate_confidence(self, stroke_pixels, total_roi_pixels, threshold):
        """Calculates stroke density and confidence level for attendance marking"""
        if total_roi_pixels <= 0:
            return "Absent", 0.0, "Low"

        density = round((stroke_pixels / total_roi_pixels) * 100, 2)

        if stroke_pixels >= threshold * 2:
            confidence = "High"
            status = "Present"
        elif stroke_pixels >= threshold:
            confidence = "Medium"
            status = "Present"
        elif stroke_pixels >= threshold * 0.5:
            confidence = "Low"
            status = "Absent"
        else:
            confidence = "High"
            status = "Absent"

        return status, density, confidence

    # =========================================================================
    # Main Attendance Processing & Analysis Pipeline
    # =========================================================================

    def process_and_analyze(
        self,
        image_path,
        students,
        session_prefix,
        signature_ratio=0.60,
        threshold_val=127,
        pixel_threshold=100,
        use_otsu=False
    ):
        """Processes attendance sheet image with deskewing, ROI thresholding,
        morphological grid line suppression, signature cropping, and confidence scoring.
        """
        img = cv2.imread(image_path)
        if img is None:
            raise ValueError(f"Unable to load image at: {image_path}")

        filename = os.path.basename(image_path)
        height, width, _ = img.shape
        num_students = len(students)

        if num_students == 0:
            raise ValueError("Student list is empty.")

        # 1. Grayscale Conversion
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # 2. Document Auto-Deskew
        img, gray, skew_angle = self._deskew_image(img, gray)

        # Save Grayscale Image
        gray_filename = f"gray_{session_prefix}_{filename}"
        gray_path = os.path.join(self.upload_folder, gray_filename)
        cv2.imwrite(gray_path, gray)

        # 3. Face Detection
        faces = []
        if self.face_cascade is not None:
            try:
                faces = self.face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4, minSize=(30, 30))
            except Exception as fe:
                print(f"[CVEngine] Face detection runtime notice: {fe}")
        
        faces_detected = len(faces)

        # 4. Preprocessing & Binarization
        blur = cv2.medianBlur(gray, 5)

        if use_otsu:
            _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        else:
            _, thresh = cv2.threshold(blur, threshold_val, 255, cv2.THRESH_BINARY_INV)

        # Save Thresholded Binarized Image
        thresh_filename = f"thresh_{session_prefix}_{filename}"
        thresh_path = os.path.join(self.upload_folder, thresh_filename)
        cv2.imwrite(thresh_path, thresh)

        # 5. Draw bounding boxes on annotated image
        annotated_img = img.copy()

        # Draw detected faces with cyan boxes
        for (fx, fy, fw, fh) in faces:
            cv2.rectangle(annotated_img, (fx, fy), (fx + fw, fy + fh), (255, 255, 0), 2)
            cv2.putText(annotated_img, "Face Detected", (fx, max(15, fy - 5)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 0), 1)

        row_height = height // num_students
        results = []

        for i, student in enumerate(students):
            y_start = i * row_height
            y_end = (i + 1) * row_height if i < num_students - 1 else height

            x_sig_start = int(width * signature_ratio)
            x_sig_end = width

            # Raw Signature ROI
            raw_sig_roi = thresh[y_start:y_end, x_sig_start:x_sig_end]
            roi_total_pixels = (y_end - y_start) * (x_sig_end - x_sig_start)

            # Morphological Grid Line & Noise Suppression
            cleaned_roi, clean_pixels = self._clean_signature_roi(raw_sig_roi)

            # Evaluate Status & Confidence
            status, density, confidence = self._calculate_confidence(clean_pixels, roi_total_pixels, pixel_threshold)

            # Crop individual signature image from original color sheet
            orig_crop = img[y_start:y_end, x_sig_start:x_sig_end]
            safe_index = re.sub(r'[^a-zA-Z0-9]', '_', student['index'])
            crop_filename = f"crop_{session_prefix}_{safe_index}.png"
            crop_path = os.path.join(self.upload_folder, crop_filename)
            cv2.imwrite(crop_path, orig_crop)

            # Draw row and signature ROI overlay
            color = (46, 204, 113) if status == "Present" else (231, 76, 60) # Green / Red (BGR)
            
            cv2.rectangle(annotated_img, (0, y_start), (width, y_end), (200, 200, 200), 1)
            cv2.rectangle(annotated_img, (x_sig_start, y_start), (x_sig_end - 2, y_end - 2), color, 2)

            label_text = f"{student['index']}: {status} ({clean_pixels}px - {confidence})"
            cv2.putText(
                annotated_img,
                label_text,
                (x_sig_start + 8, y_start + max(20, row_height // 2)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.50,
                color,
                2,
                cv2.LINE_AA
            )

            results.append({
                "index": student["index"],
                "name": student["name"],
                "status": status,
                "pixel_count": clean_pixels,
                "density_percentage": density,
                "confidence": confidence,
                "crop_image": f"uploads/{crop_filename}"
            })

        # Save Annotated Image
        annotated_filename = f"annotated_{session_prefix}_{filename}"
        annotated_path = os.path.join(self.upload_folder, annotated_filename)
        cv2.imwrite(annotated_path, annotated_img)

        step_images = {
            "original": f"uploads/{filename}",
            "annotated": f"uploads/{annotated_filename}",
            "grayscale": f"uploads/{gray_filename}",
            "binarized": f"uploads/{thresh_filename}"
        }

        return step_images, results, faces_detected

    # =========================================================================
    # PDF & Excel Export Generators
    # =========================================================================

    def generate_pdf_report(self, session, records):
        """Generates a styled PDF report for an attendance session"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1e293b'), spaceAfter=8)
        meta_style = ParagraphStyle('MetaStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=14)

        elements = []

        # Header Title
        elements.append(Paragraph(f"SAMS Attendance Report - {session['title']}", title_style))
        rate = (session['present_count'] / session['total_students'] * 100) if session['total_students'] > 0 else 0
        meta_text = f"<b>Session Key:</b> {session['session_key']} &nbsp;|&nbsp; <b>Date:</b> {session['date_str']} &nbsp;|&nbsp; <b>Present:</b> {session['present_count']}/{session['total_students']} ({rate:.1f}%)"
        elements.append(Paragraph(meta_text, meta_style))
        elements.append(Spacer(1, 10))

        # Table Header & Rows
        table_data = [["Index", "Student Name", "Status", "Pixel Density", "Signature Crop"]]

        for r in records:
            status_color = colors.HexColor('#22c55e') if r['status'] == 'Present' else colors.HexColor('#ef4444')
            status_p = Paragraph(f"<font color='{status_color}'><b>{r['status']}</b></font>", styles['Normal'])
            
            crop_cell = "N/A"
            if r['crop_image']:
                crop_full_path = os.path.join(os.path.dirname(self.upload_folder), r['crop_image'])
                if os.path.exists(crop_full_path):
                    try:
                        crop_cell = RLImage(crop_full_path, width=80, height=28)
                    except Exception:
                        crop_cell = "Img Err"

            table_data.append([
                r['student_index'],
                r['student_name'],
                status_p,
                f"{r['pixel_count']} px",
                crop_cell
            ])

        t = Table(table_data, colWidths=[90, 180, 80, 80, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))

        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel_report(self, session, records):
        """Generates a styled Excel (.xlsx) report for an attendance session"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"

        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        
        present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        present_font = Font(name="Calibri", size=11, bold=True, color="15803D")
        
        absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        absent_font = Font(name="Calibri", size=11, bold=True, color="B91C1C")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Title Block
        ws.merge_cells("A1:E1")
        ws["A1"] = f"SAMS Attendance Report - {session['title']}"
        ws["A1"].font = title_font

        ws["A2"] = f"Session Key: {session['session_key']} | Date: {session['date_str']}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")

        # Table Header
        headers = ["Index Number", "Student Name", "Status", "Raw Pixel Count", "Manually Overridden"]
        ws.append([]) # Row 3 blank
        ws.append(headers) # Row 4

        for col_idx in range(1, 6):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Table Data Rows
        for row_idx, r in enumerate(records, start=5):
            ws.cell(row=row_idx, column=1, value=r['student_index']).border = thin_border
            ws.cell(row=row_idx, column=2, value=r['student_name']).border = thin_border
            
            status_cell = ws.cell(row=row_idx, column=3, value=r['status'])
            status_cell.border = thin_border
            status_cell.alignment = Alignment(horizontal="center")

            if r['status'] == 'Present':
                status_cell.fill = present_fill
                status_cell.font = present_font
            else:
                status_cell.fill = absent_fill
                status_cell.font = absent_font

            ws.cell(row=row_idx, column=4, value=r['pixel_count']).border = thin_border
            ws.cell(row=row_idx, column=5, value="Yes" if r['is_manually_overridden'] else "No").border = thin_border

        # Adjust Column Widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
